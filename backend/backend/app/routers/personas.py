from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from pydantic import BaseModel
from typing import Optional
import openpyxl, io
from ..database import get_db, Persona
from ..auth import get_current_user, Usuario

router = APIRouter(prefix="/api/personas", tags=["personas"])


class PersonaCreate(BaseModel):
    tipo:         str
    nombre:       str
    email:        str
    programa:     Optional[str] = None
    matricula:    Optional[str] = None
    cargo:        Optional[str] = None
    departamento: Optional[str] = None


def persona_dict(p: Persona):
    return {
        "id":           p.id,
        "tipo":         p.tipo,
        "nombre":       p.nombre,
        "email":        p.email,
        "programa":     p.programa,
        "matricula":    p.matricula,
        "cargo":        p.cargo,
        "departamento": p.departamento,
        "created_at":   p.created_at.isoformat() if p.created_at else None,
    }


@router.get("/")
async def listar_personas(
    tipo:     Optional[str] = Query(None),
    busqueda: Optional[str] = Query(None),
    programa: Optional[str] = Query(None),
    limit:    int           = Query(500),
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(get_current_user),
):
    q = select(Persona)
    filters = []
    if tipo:     filters.append(Persona.tipo == tipo)
    if programa: filters.append(Persona.programa == programa)
    if busqueda:
        filters.append(Persona.nombre.ilike(f"%{busqueda}%"))
    if filters:
        q = q.where(and_(*filters))
    q = q.order_by(Persona.nombre).limit(limit)
    result = await db.execute(q)
    return [persona_dict(p) for p in result.scalars().all()]


@router.post("/", status_code=201)
async def crear_persona(
    data: PersonaCreate,
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(get_current_user),
):
    p = Persona(**data.model_dump())
    db.add(p)
    await db.commit()
    await db.refresh(p)
    return persona_dict(p)


@router.delete("/{persona_id}")
async def eliminar_persona(
    persona_id: int,
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(get_current_user),
):
    result = await db.execute(select(Persona).where(Persona.id == persona_id))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(status_code=404, detail="Persona no encontrada")
    await db.delete(p)
    await db.commit()
    return {"ok": True}


@router.post("/importar-excel")
async def importar_excel(
    tipo: str,
    archivo: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(get_current_user),
):
    content = await archivo.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    creados = 0
    omitidos = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        fila = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row) if i < len(headers)}
        nombre = fila.get("nombre", "")
        email  = fila.get("email", "")
        if not nombre or not email:
            omitidos += 1
            continue

        p = Persona(
            tipo         = tipo,
            nombre       = nombre,
            email        = email,
            programa     = fila.get("programa") or None,
            matricula    = fila.get("matricula") or None,
            cargo        = fila.get("cargo") or None,
            departamento = fila.get("departamento") or None,
        )
        db.add(p)
        creados += 1

    await db.commit()
    return {"creados": creados, "omitidos": omitidos}
