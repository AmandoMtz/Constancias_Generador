from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from pathlib import Path
from uuid import uuid4
from openpyxl import load_workbook
from typing import Optional
from ..auth import get_current_user, Usuario
from ..config import UPLOADS_EXCEL_DIR

router = APIRouter(prefix="/api/excel", tags=["excel"])

UPLOAD_DIR = UPLOADS_EXCEL_DIR


def _excel_path(excel_id: str) -> Optional[Path]:
    posibles = sorted(UPLOAD_DIR.glob(f"{excel_id}.*"))
    return posibles[0] if posibles else None


def _sheet_columns(ws):
    headers = []
    if ws.max_row < 1:
        return headers
    for idx, cell in enumerate(ws[1], start=1):
        valor = cell.value
        headers.append(str(valor).strip() if valor not in (None, "") else f"col_{idx}")
    return headers


def _sheet_info(ws):
    cols = _sheet_columns(ws)
    total_rows = max(ws.max_row - 1, 0)
    return {"name": ws.title, "columns": cols, "total_rows": total_rows}


def _preview_rows(ws, limit=10):
    headers = _sheet_columns(ws)
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, limit + 1), values_only=True):
        item = {}
        for idx, header in enumerate(headers):
            val = row[idx] if idx < len(row) else None
            item[header] = "" if val is None else str(val)
        rows.append(item)
    return rows


@router.post('/upload')
async def upload_excel(
    archivo: UploadFile = File(...),
    _: Usuario = Depends(get_current_user),
):
    if not archivo.filename.lower().endswith('.xlsx'):
        raise HTTPException(status_code=400, detail='Solo se permiten archivos .xlsx')
    excel_id = uuid4().hex
    destino = UPLOAD_DIR / f"{excel_id}.xlsx"
    destino.write_bytes(await archivo.read())

    try:
        wb = load_workbook(destino, read_only=True, data_only=True)
        sheets = [_sheet_info(ws) for ws in wb.worksheets]
        wb.close()
    except Exception as e:
        destino.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail=f'No se pudo leer el Excel: {e}')

    return {
        'excel_id': excel_id,
        'filename': archivo.filename,
        'sheets': sheets,
    }


@router.get('/{excel_id}/sheets')
async def listar_sheets(excel_id: str, _: Usuario = Depends(get_current_user)):
    ruta = _excel_path(excel_id)
    if not ruta:
        raise HTTPException(status_code=404, detail='Excel no encontrado')
    wb = load_workbook(ruta, read_only=True, data_only=True)
    sheets = [_sheet_info(ws) for ws in wb.worksheets]
    wb.close()
    return {'excel_id': excel_id, 'sheets': sheets}


@router.get('/{excel_id}/columns')
async def columnas_sheet(
    excel_id: str,
    sheet_name: str = Query(...),
    _: Usuario = Depends(get_current_user),
):
    ruta = _excel_path(excel_id)
    if not ruta:
        raise HTTPException(status_code=404, detail='Excel no encontrado')
    wb = load_workbook(ruta, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise HTTPException(status_code=404, detail='Hoja no encontrada')
    ws = wb[sheet_name]
    data = _sheet_info(ws)
    wb.close()
    return data


@router.get('/{excel_id}/preview')
async def preview_sheet(
    excel_id: str,
    sheet_name: str = Query(...),
    limit: int = Query(10, ge=1, le=25),
    _: Usuario = Depends(get_current_user),
):
    ruta = _excel_path(excel_id)
    if not ruta:
        raise HTTPException(status_code=404, detail='Excel no encontrado')
    wb = load_workbook(ruta, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise HTTPException(status_code=404, detail='Hoja no encontrada')
    ws = wb[sheet_name]
    payload = {
        'name': ws.title,
        'columns': _sheet_columns(ws),
        'total_rows': max(ws.max_row - 1, 0),
        'rows': _preview_rows(ws, limit=limit),
    }
    wb.close()
    return payload
